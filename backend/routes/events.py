# -*- coding: utf-8 -*-
"""事件 CRUD 路由 + 导入导出"""

import re
import calendar as cal_mod
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from flask import Blueprint, request, jsonify, send_file
from ..database import get_db
from ..auth import require_login, get_current_user_id

events_bp = Blueprint('events', __name__)


def _event_row(r):
    return {
        'id': r['id'], 'title': r['title'], 'date': r['date'],
        'time': r['time'], 'color': r['color'], 'completed': bool(r['completed']),
    }


def _require_uid():
    uid = get_current_user_id()
    if not uid:
        from flask import abort
        abort(401)
    return uid


@events_bp.route('/api/events', methods=['GET'])
def get_events():
    uid = _require_uid()
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if year and month:
        start_date = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1:04d}-01-01"
        else:
            end_date = f"{year:04d}-{month + 1:02d}-01"
    else:
        today = datetime.now()
        year, month = today.year, today.month
        start_date = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1:04d}-01-01"
        else:
            end_date = f"{year:04d}-{month + 1:02d}-01"

    db = get_db()
    rows = db.execute(
        'SELECT * FROM events WHERE user_id = ? AND date >= ? AND date < ? ORDER BY date, time',
        (uid, start_date, end_date)
    ).fetchall()
    return jsonify([_event_row(r) for r in rows])


@events_bp.route('/api/events/day', methods=['GET'])
def get_day_events():
    uid = _require_uid()
    date_str = request.args.get('date', '')
    if not date_str:
        date_str = datetime.now().strftime('%Y-%m-%d')
    db = get_db()
    rows = db.execute(
        'SELECT * FROM events WHERE user_id = ? AND date = ? ORDER BY time',
        (uid, date_str)
    ).fetchall()
    return jsonify([_event_row(r) for r in rows])


@events_bp.route('/api/events/week', methods=['GET'])
def get_week_events():
    uid = _require_uid()
    date_str = request.args.get('date', '')
    if date_str:
        d = datetime.strptime(date_str, '%Y-%m-%d')
    else:
        d = datetime.now()
    monday = d - timedelta(days=d.weekday())
    sunday = monday + timedelta(days=6)
    db = get_db()
    rows = db.execute(
        'SELECT * FROM events WHERE user_id = ? AND date >= ? AND date <= ? ORDER BY date, time',
        (uid, monday.strftime('%Y-%m-%d'), sunday.strftime('%Y-%m-%d'))
    ).fetchall()
    return jsonify([_event_row(r) for r in rows])


@events_bp.route('/api/events', methods=['POST'])
@require_login
def create_event():
    data = request.get_json()
    title = data.get('title', '').strip()
    if not title:
        return jsonify({'error': '事件标题不能为空'}), 400
    date_str = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    time_str = data.get('time', '')
    color = data.get('color', '#4A90D9')
    completed = data.get('completed', False)
    uid = get_current_user_id()

    db = get_db()
    cur = db.execute(
        'INSERT INTO events (user_id, title, date, time, color, completed) VALUES (?, ?, ?, ?, ?, ?)',
        (uid, title, date_str, time_str, color, 1 if completed else 0)
    )
    db.commit()
    return jsonify({
        'id': cur.lastrowid, 'title': title, 'date': date_str,
        'time': time_str, 'color': color, 'completed': completed,
    }), 201


@events_bp.route('/api/events/<int:event_id>', methods=['PUT'])
@require_login
def update_event(event_id):
    data = request.get_json()
    uid = get_current_user_id()
    db = get_db()
    existing = db.execute(
        'SELECT * FROM events WHERE id = ? AND user_id = ?', (event_id, uid)
    ).fetchone()
    if not existing:
        return jsonify({'error': '事件不存在'}), 404

    title = data.get('title', existing['title']).strip()
    date_str = data.get('date', existing['date'])
    time_str = data.get('time', existing['time'])
    color = data.get('color', existing['color'])
    completed = data.get('completed', existing['completed'])
    if isinstance(completed, bool):
        completed = 1 if completed else 0

    db.execute(
        'UPDATE events SET title=?, date=?, time=?, color=?, completed=?, updated_at=? WHERE id=?',
        (title, date_str, time_str, color, completed,
         datetime.now().strftime('%Y-%m-%d %H:%M:%S'), event_id)
    )
    db.commit()
    return jsonify({'success': True})


@events_bp.route('/api/events/<int:event_id>/toggle', methods=['POST'])
@require_login
def toggle_event(event_id):
    uid = get_current_user_id()
    db = get_db()
    existing = db.execute(
        'SELECT completed FROM events WHERE id = ? AND user_id = ?', (event_id, uid)
    ).fetchone()
    if not existing:
        return jsonify({'error': '事件不存在'}), 404
    new_status = 0 if existing['completed'] else 1
    db.execute(
        'UPDATE events SET completed=?, updated_at=? WHERE id=?',
        (new_status, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), event_id)
    )
    db.commit()
    return jsonify({'completed': bool(new_status)})


@events_bp.route('/api/events/<int:event_id>', methods=['DELETE'])
@require_login
def delete_event(event_id):
    uid = get_current_user_id()
    db = get_db()
    db.execute('DELETE FROM events WHERE id = ? AND user_id = ?', (event_id, uid))
    db.commit()
    return jsonify({'success': True})


# ========== 导入导出 ==========

WEEKDAYS_CN = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']

# 样式常量
_THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
_HEADER_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
_HEADER_FONT = Font(name='微软雅黑', bold=True, size=11)
_MONTH_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
_MONTH_FONT = Font(name='微软雅黑', bold=True, size=11, color='C55A11')
_DAY_FONT = Font(name='微软雅黑', size=10, color='808080')
_EVENT_FONT = Font(name='微软雅黑', size=10)
_CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)


def _build_calendar_data(uid, year, month):
    """构建单月日历数据：{(row_idx, col_idx): [event_dict, ...]}, day_map"""
    db = get_db()
    start = f'{year:04d}-{month:02d}-01'
    if month == 12:
        end = f'{year + 1:04d}-01-01'
    else:
        end = f'{year:04d}-{month + 1:02d}-01'
    rows = db.execute(
        'SELECT date, title, color, completed FROM events WHERE user_id=? AND date>=? AND date<? ORDER BY date, time',
        (uid, start, end)).fetchall()

    # 按日期分组事件（保留完整信息）
    day_events = {}
    for r in rows:
        day_str = r['date']
        day_events.setdefault(day_str, []).append({
            'title': r['title'],
            'color': r['color'] or '#4A90D9',
            'completed': bool(r['completed']),
        })

    # 生成日历布局
    first_day = datetime(year, month, 1)
    first_weekday = first_day.weekday()  # 0=Mon
    total_days = cal_mod.monthrange(year, month)[1]

    # col_idx: 1=周一 ... 7=周日
    # 第一周的 row_idx = 2（第1行=标题, 第2行=月份标记）
    calendar_cells = {}
    current_row = 1  # 第1行：星期标题（=Excel row 2，因为 row 1 是月份行）
    col = first_weekday + 1
    for day in range(1, total_days + 1):
        date_str = f'{year:04d}-{month:02d}-{day:02d}'
        events = day_events.get(date_str, [])
        calendar_cells[(current_row, col)] = (day, events)
        col += 1
        if col > 7:
            col = 1
            current_row += 1

    return calendar_cells, current_row


def _style_cell(cell, font, fill=None, alignment=None, border=_THIN_BORDER):
    cell.font = font
    cell.alignment = alignment or _LEFT_ALIGN
    cell.border = border
    if fill:
        cell.fill = fill


@events_bp.route('/api/events/export', methods=['GET'])
@require_login
def export_events():
    uid = get_current_user_id()
    export_all = request.args.get('all', '0') == '1'

    # 解析 year 和 month —— 允许缺失以区分"按年导出"和"按月导出"
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    try:
        if export_all:
            return _do_export_all(uid)
        if year and month:
            return _do_export(uid, year, month)
        if year and not month:
            # 仅传了 year，未传 month → 按年导出（12 个月分 Sheet）
            return _do_export_year(uid, year)
        # 未传任何参数 → 默认导出当前月
        now = datetime.now()
        return _do_export(uid, now.year, now.month)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'导出失败: {str(e)}'}), 500


@events_bp.route('/api/events/years', methods=['GET'])
@require_login
def get_available_years():
    """获取当前用户有事件数据的年份列表，供前端导出选择"""
    uid = get_current_user_id()
    db = get_db()
    rows = db.execute(
        'SELECT DISTINCT substr(date,1,4) as yr '
        'FROM events WHERE user_id=? AND date IS NOT NULL ORDER BY yr',
        (uid,)).fetchall()
    years = [int(r['yr']) for r in rows if r['yr']]
    # 确保当前年份始终在列表中
    current_year = datetime.now().year
    if current_year not in years:
        years.append(current_year)
    years.sort()
    return jsonify({'years': years})


def _write_month_sheet(wb, uid, year, month, sheet_name=None):
    """写入单月数据到一个 worksheet，返回该 sheet"""
    ws = wb.create_sheet(title=sheet_name or f'{year}.{month}')
    # 列宽
    ws.column_dimensions['A'].width = 4
    for i in range(2, 9):
        ws.column_dimensions[get_column_letter(i)].width = 22

    # 第1行：月份标记
    ws.merge_cells('A1:H1')
    cell = ws.cell(1, 1, f'{year}.{month}')
    _style_cell(cell, _MONTH_FONT, _MONTH_FILL, _CENTER_ALIGN)

    # 第2行：星期标题
    ws.cell(2, 1).border = _THIN_BORDER
    for i, wd in enumerate(WEEKDAYS_CN):
        cell = ws.cell(2, i + 2, wd)
        _style_cell(cell, _HEADER_FONT, _HEADER_FILL, _CENTER_ALIGN)

    # 获取当月数据并写入
    cells, max_row = _build_calendar_data(uid, year, month)
    for r in range(3, max_row + 4):  # 留一些空行
        for c in range(1, 9):
            cell = ws.cell(r, c)
            if c == 1:
                cell.border = _THIN_BORDER
                continue
            key = (r - 2, c - 1)
            if key in cells:
                day, events = cells[key]
                if events:
                    from openpyxl.cell.rich_text import CellRichText, TextBlock
                    from openpyxl.cell.text import InlineFont

                    day_font = InlineFont(rFont='微软雅黑', sz='10', color='808080')
                    blocks = [TextBlock(day_font, f'{day}\n')]

                    for i, ev in enumerate(events):
                        color_val = (ev['color'] or '#4A90D9').lstrip('#')
                        ev_font = InlineFont(
                            rFont='微软雅黑', sz='10',
                            color=color_val,
                            strike=ev['completed'],
                        )
                        prefix = '• ' if i == 0 else '\n• '
                        blocks.append(TextBlock(ev_font, f'{prefix}{ev["title"]}'))

                    cell.value = CellRichText(blocks)
                    cell.alignment = _LEFT_ALIGN
                    cell.border = _THIN_BORDER
                else:
                    cell.value = day
                    _style_cell(cell, _DAY_FONT, alignment=_CENTER_ALIGN)
            else:
                _style_cell(cell, _EVENT_FONT)

    ws.freeze_panes = 'B3'
    return ws


def _do_export(uid, year, month):
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)
    _write_month_sheet(wb, uid, year, month, sheet_name='日历清单')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'calendar-{year:04d}{month:02d}.xlsx')


def _do_export_year(uid, year):
    """导出指定年份的所有月份数据，按月份分 sheet"""
    db = get_db()
    start = f'{year:04d}-01-01'
    end = f'{year + 1:04d}-01-01'
    rows = db.execute(
        'SELECT DISTINCT substr(date,6,2) as mo '
        'FROM events WHERE user_id=? AND date>=? AND date<? ORDER BY mo',
        (uid, start, end)).fetchall()

    if not rows:
        return jsonify({'error': f'{year}年没有可导出的数据'}), 400

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for r in rows:
        mo = int(r['mo'])
        _write_month_sheet(wb, uid, year, mo, sheet_name=f'{year}.{mo}')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'calendar-{year:04d}.xlsx')


def _do_export_all(uid):
    """导出当前用户所有事件数据，按月份分 sheet"""
    db = get_db()
    rows = db.execute(
        'SELECT DISTINCT substr(date,1,4) as yr, substr(date,6,2) as mo '
        'FROM events WHERE user_id=? ORDER BY yr, mo', (uid,)
    ).fetchall()

    if not rows:
        return jsonify({'error': '没有可导出的数据'}), 400

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for r in rows:
        yr = int(r['yr'])
        mo = int(r['mo'])
        _write_month_sheet(wb, uid, yr, mo, sheet_name=f'{yr}.{mo}')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='calendar-all.xlsx')


@events_bp.route('/api/events/import', methods=['POST'])
@require_login
def import_events():
    uid = get_current_user_id()
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请选择文件'}), 400

    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()), rich_text=True)
    except Exception:
        return jsonify({'error': '文件格式不正确，请使用 Excel (.xlsx) 文件'}), 400

    ws = wb.active
    events = []  # list of (date_str, title, color, completed)
    current_year = None
    current_month = None

    for row in ws.iter_rows(min_row=1):
        if not row or len(row) < 2:
            continue
        a_cell = row[0]
        a_val = a_cell.value
        # 月份标记：支持数字 (2026.8) 和字符串 ('2026.8') 两种格式
        a_str = str(a_val).strip() if a_val is not None else ''
        m_month = re.match(r'^(\d{4})\.(\d{1,2})$', a_str)
        if m_month:
            current_year = int(m_month.group(1))
            current_month = int(m_month.group(2))
            continue
        # 标题行
        b_val = row[1].value
        if str(b_val).strip() == '周一':
            continue
        if current_year is None or current_month is None:
            continue

        for col_idx in range(1, 8):
            if col_idx >= len(row):
                break
            cell = row[col_idx]
            cell_val = cell.value
            if cell_val is None:
                continue
            cell_str = str(cell_val).strip()
            m = re.match(r'(\d+)\s*\n?\s*[•\-\s]*(.*)', cell_str, re.DOTALL)
            if not m:
                continue
            day = int(m.group(1))
            desc_raw = m.group(2).strip()
            if day < 1 or day > 31 or not desc_raw:
                continue
            date_str = f'{current_year:04d}-{current_month:02d}-{day:02d}'

            # 解析每个事件的文本、颜色、删除线
            # 支持 CellRichText（多 run）和普通字符串
            event_texts = _parse_cell_events(cell, desc_raw)
            for title, color, completed in event_texts:
                title = re.sub(r'\s+', ' ', title).strip()
                if title and len(title) > 1:
                    events.append((date_str, title[:256], color, completed))

    wb.close()

    if not events:
        return jsonify({'success': True, 'inserted': 0, 'skipped': 0, 'message': '文件中未检测到有效事件数据'})

    db = get_db()

    # 批量查重：一次查询获取所有已存在的 (date, title) 对
    inserted = 0
    skipped = 0
    to_insert = []
    # 分批查重（每批 50 条，避免 IN 子句过长）
    batch_size = 50
    for i in range(0, len(events), batch_size):
        batch = events[i:i + batch_size]
        # 构建 IN 子句
        placeholders = ','.join(['(?,?)'] * len(batch))
        flat_params = []
        for date_str, title, _color, _completed in batch:
            flat_params.extend([date_str, title])
        existing_rows = db.execute(
            f'SELECT date, title FROM events WHERE user_id=? AND (date, title) IN ({placeholders})',
            [uid] + flat_params).fetchall()
        existing_set = set((r['date'], r['title']) for r in existing_rows)

        for date_str, title, color, completed in batch:
            if (date_str, title) in existing_set:
                skipped += 1
                continue
            to_insert.append((uid, title, date_str, '', color, 1 if completed else 0))
            inserted += 1

    # 批量插入
    if to_insert:
        db.executemany(
            'INSERT INTO events (user_id, title, date, time, color, completed) VALUES (?, ?, ?, ?, ?, ?)',
            to_insert)
        db.commit()

    return jsonify({
        'success': True, 'inserted': inserted, 'skipped': skipped,
        'message': f'导入完成：新增 {inserted} 条，跳过重复 {skipped} 条'
    })


def _parse_cell_events(cell, desc_raw):
    """从单元格中解析事件列表，提取每个事件的标题、颜色和删除线状态。
    
    支持 CellRichText（多 run）和普通字符串。
    返回: [(title, color_hex, completed_bool), ...]
    """
    from openpyxl.cell.rich_text import CellRichText

    # 先按行分割描述文本
    lines = re.split(r'[\n]+', desc_raw)
    # 去掉每行前缀的 • 或 -
    lines = [re.sub(r'^[•\-\s]+', '', line).strip() for line in lines]
    lines = [line for line in lines if line]

    if not lines:
        return []

    # 尝试从 CellRichText 中提取每个 run 的颜色和删除线
    rich_info = []  # [(text, color, strike), ...]
    if isinstance(cell.value, CellRichText):
        for block in cell.value:
            if hasattr(block, 'text') and block.text:
                text = block.text
                font = block.font
                color = '#4A90D9'  # 默认颜色
                strike = False
                if font and font.color and font.color.rgb:
                    rgb = str(font.color.rgb)
                    # rgb 格式可能是 'FF4A90D9' 或 '4A90D9'
                    if len(rgb) == 8:
                        color = '#' + rgb[2:]
                    elif len(rgb) == 6:
                        color = '#' + rgb
                if font and font.strike:
                    strike = True
                rich_info.append((text, color, strike))

    if not rich_info:
        # 普通 String 单元格，检查整个单元格的字体样式
        font = cell.font
        color = '#4A90D9'
        strike = False
        if font and font.color and font.color.rgb:
            rgb = str(font.color.rgb)
            if len(rgb) == 8:
                color = '#' + rgb[2:]
            elif len(rgb) == 6:
                color = '#' + rgb
        if font and font.strike:
            strike = True
        return [(line, color, strike) for line in lines]

    # CellRichText: 将 rich_info 的 run 与分割后的事件行匹配
    # 将所有 rich text 拼接，然后按事件分割
    full_text = ''.join(t for t, _c, _s in rich_info)

    # 尝试按位置匹配每个事件
    results = []
    search_pos = 0
    for line in lines:
        pos = full_text.find(line, search_pos)
        if pos >= 0:
            # 找到该事件文本在 rich_text 中的范围，检查覆盖它的 run
            event_end = pos + len(line)
            event_color = '#4A90D9'
            event_strike = False
            char_pos = 0
            for text, color, strike in rich_info:
                run_start = char_pos
                run_end = char_pos + len(text)
                # 如果 run 与事件文本有重叠
                if run_start < event_end and run_end > pos:
                    event_color = color
                    event_strike = strike
                    break
                char_pos = run_end
            search_pos = event_end
            results.append((line, event_color, event_strike))
        else:
            results.append((line, '#4A90D9', False))

    return results


@events_bp.route('/api/events/template', methods=['GET'])
def download_template():
    """动态生成空白导入模板"""
    now = datetime.now()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '日历清单'

    # 列宽
    ws.column_dimensions['A'].width = 4
    for i in range(2, 9):
        ws.column_dimensions[get_column_letter(i)].width = 22

    # 第1行：月份标记
    ws.merge_cells('A1:H1')
    cell = ws.cell(1, 1, f'{now.year}.{now.month}')
    _style_cell(cell, _MONTH_FONT, _MONTH_FILL, _CENTER_ALIGN)

    # 第2行：星期标题
    ws.cell(2, 1).border = _THIN_BORDER
    for i, wd in enumerate(WEEKDAYS_CN):
        cell = ws.cell(2, i + 2, wd)
        _style_cell(cell, _HEADER_FONT, _HEADER_FILL, _CENTER_ALIGN)

    # 生成当月日历布局（空事件，不查数据库）
    first_day = datetime(now.year, now.month, 1)
    first_weekday = first_day.weekday()
    total_days = cal_mod.monthrange(now.year, now.month)[1]

    current_row = 1
    col = first_weekday + 1
    day_cells = {}
    for day in range(1, total_days + 1):
        day_cells[(current_row, col)] = day
        col += 1
        if col > 7:
            col = 1
            current_row += 1

    for r in range(3, current_row + 3):
        for c in range(1, 9):
            cell = ws.cell(r, c)
            if c == 1:
                cell.border = _THIN_BORDER
                continue
            key = (r - 2, c - 1)
            if key in day_cells:
                cell.value = day_cells[key]
                _style_cell(cell, _DAY_FONT, alignment=_CENTER_ALIGN)
            else:
                _style_cell(cell, _EVENT_FONT)

    ws.freeze_panes = 'B3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='calendar-template.xlsx')
